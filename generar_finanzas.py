# -*- coding: utf-8 -*-
"""
Genera 'Finanzas_Personales.xlsx' con 4 hojas:
  1. Dashboard (Resumen General)
  2. Ingresos
  3. Gastos
  4. Metas de Ahorro
Incluye: formulas automaticas, menus desplegables, semaforos (formato
condicional) y graficos. Moneda: Cordoba nicaraguense (C$).
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import date

# ---------------------------------------------------------------------------
# Configuracion general / paleta de colores
# ---------------------------------------------------------------------------
MONEDA = '"C$" #,##0.00'
PORC = '0.0%'

AZUL_OSC   = "1F3864"   # encabezados principales
AZUL       = "2E5496"
AZUL_CLARO = "D9E1F2"
GRIS_CLARO = "F2F2F2"
VERDE      = "C6EFCE"; VERDE_TXT  = "006100"
AMARILLO   = "FFEB9C"; AMAR_TXT   = "9C6500"
ROJO       = "FFC7CE"; ROJO_TXT   = "9C0006"
BLANCO     = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDE = Border(left=thin, right=thin, top=thin, bottom=thin)

CATEGORIAS = ["Alimentación", "Transporte", "Servicios", "Estudios",
              "Casa", "Salud", "Entretenimiento", "Otros"]
MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
         "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

MAXFILA = 500  # rango de filas para registros

def d_(s):
    """Convierte 'YYYY-MM-DD' en un objeto date real."""
    y, m, dd = map(int, s.split("-"))
    return date(y, m, dd)

wb = Workbook()

# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------
def titulo(ws, celda, texto, size=16):
    c = ws[celda]
    c.value = texto
    c.font = Font(bold=True, size=size, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal="left", vertical="center")

def encabezado(c, texto):
    c.value = texto
    c.font = Font(bold=True, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDE

# ===========================================================================
# HOJA 2: INGRESOS
# ===========================================================================
wi = wb.active
wi.title = "Ingresos"
titulo(wi, "A1", "REGISTRO DE INGRESOS")
wi.merge_cells("A1:C1")
wi.row_dimensions[1].height = 26

for col, txt in zip("ABC", ["Fecha", "Descripción", "Monto (C$)"]):
    encabezado(wi[f"{col}2"], txt)

# Helpers (Año y Mes) en columnas E y F -> usadas por el Dashboard
encabezado(wi["E2"], "Año")
encabezado(wi["F2"], "Mes")

ingresos_demo = [
    ("2026-06-01", "Salario", 25000),
    ("2026-06-05", "Venta freelance", 4500),
    ("2026-06-15", "Reembolso", 1200),
    ("2026-05-01", "Salario", 25000),
    ("2026-05-10", "Bono", 3000),
]
for i, (f, d, m) in enumerate(ingresos_demo, start=3):
    wi[f"A{i}"] = d_(f)
    wi[f"B{i}"] = d
    wi[f"C{i}"] = m

for r in range(3, MAXFILA + 1):
    wi[f"A{r}"].number_format = "yyyy-mm-dd"
    wi[f"C{r}"].number_format = MONEDA
    wi[f"E{r}"] = f'=IF(A{r}="","",YEAR(A{r}))'
    wi[f"F{r}"] = f'=IF(A{r}="","",MONTH(A{r}))'
    for col in "ABC":
        wi[f"{col}{r}"].border = BORDE

# Total
wi["B1"]  # noop
wi.column_dimensions["A"].width = 14
wi.column_dimensions["B"].width = 32
wi.column_dimensions["C"].width = 16
wi.column_dimensions["E"].width = 8
wi.column_dimensions["F"].width = 8
wi.sheet_view.showGridLines = False

# ===========================================================================
# HOJA 3: GASTOS
# ===========================================================================
wg = wb.create_sheet("Gastos")
titulo(wg, "A1", "REGISTRO DE GASTOS")
wg.merge_cells("A1:D1")
wg.row_dimensions[1].height = 26

for col, txt in zip("ABCD", ["Fecha", "Categoría", "Descripción", "Monto (C$)"]):
    encabezado(wg[f"{col}2"], txt)
encabezado(wg["F2"], "Año")
encabezado(wg["G2"], "Mes")

gastos_demo = [
    ("2026-06-02", "Alimentación", "Supermercado", 3800),
    ("2026-06-03", "Transporte", "Combustible", 1500),
    ("2026-06-04", "Servicios", "Luz y agua", 1800),
    ("2026-06-06", "Estudios", "Curso online", 1200),
    ("2026-06-08", "Casa", "Reparación", 2500),
    ("2026-06-10", "Salud", "Farmacia", 900),
    ("2026-06-12", "Entretenimiento", "Cine y salida", 1100),
    ("2026-06-14", "Alimentación", "Restaurante", 1400),
    ("2026-05-03", "Alimentación", "Supermercado", 3600),
    ("2026-05-07", "Transporte", "Bus", 800),
]
for i, (f, cat, d, m) in enumerate(gastos_demo, start=3):
    wg[f"A{i}"] = d_(f)
    wg[f"B{i}"] = cat
    wg[f"C{i}"] = d
    wg[f"D{i}"] = m

for r in range(3, MAXFILA + 1):
    wg[f"A{r}"].number_format = "yyyy-mm-dd"
    wg[f"D{r}"].number_format = MONEDA
    wg[f"F{r}"] = f'=IF(A{r}="","",YEAR(A{r}))'
    wg[f"G{r}"] = f'=IF(A{r}="","",MONTH(A{r}))'
    for col in "ABCD":
        wg[f"{col}{r}"].border = BORDE

# Menu desplegable de categorias
dv = DataValidation(
    type="list",
    formula1='"' + ",".join(CATEGORIAS) + '"',
    allow_blank=True,
    showDropDown=False,
)
dv.error = "Selecciona una categoría de la lista."
dv.errorTitle = "Categoría inválida"
dv.prompt = "Elige una categoría"
dv.promptTitle = "Categoría"
wg.add_data_validation(dv)
dv.add(f"B3:B{MAXFILA}")

wg.column_dimensions["A"].width = 14
wg.column_dimensions["B"].width = 18
wg.column_dimensions["C"].width = 30
wg.column_dimensions["D"].width = 16
wg.column_dimensions["F"].width = 8
wg.column_dimensions["G"].width = 8
wg.sheet_view.showGridLines = False

# ===========================================================================
# HOJA 4: METAS DE AHORRO
# ===========================================================================
wm = wb.create_sheet("Metas")
titulo(wm, "A1", "METAS DE AHORRO")
wm.merge_cells("A1:E1")
wm.row_dimensions[1].height = 26

for col, txt in zip("ABCDE",
                    ["Meta", "Objetivo (C$)", "Ahorrado (C$)",
                     "Falta (C$)", "% Avance"]):
    encabezado(wm[f"{col}2"], txt)

metas_demo = [
    ("Fondo para la casa", 150000, 35000),
    ("Fondo de emergencia", 50000, 18000),
    ("Negocio propio", 80000, 12000),
    ("Vacaciones", 20000, 6000),
]
for i, (meta, obj, ahorr) in enumerate(metas_demo, start=3):
    wm[f"A{i}"] = meta
    wm[f"B{i}"] = obj
    wm[f"C{i}"] = ahorr

ULTIMA_META = 2 + 12  # permitir hasta 12 metas
for r in range(3, ULTIMA_META + 1):
    wm[f"B{r}"].number_format = MONEDA
    wm[f"C{r}"].number_format = MONEDA
    wm[f"D{r}"] = f'=IF(B{r}="","",MAX(0,B{r}-C{r}))'
    wm[f"D{r}"].number_format = MONEDA
    wm[f"E{r}"] = f'=IF(B{r}="","",IF(B{r}=0,0,C{r}/B{r}))'
    wm[f"E{r}"].number_format = PORC
    for col in "ABCDE":
        wm[f"{col}{r}"].border = BORDE

# Semaforo sobre el % de avance de cada meta
rango_avance = f"E3:E{ULTIMA_META}"
wm.conditional_formatting.add(rango_avance,
    CellIsRule(operator="greaterThanOrEqual", formula=["0.75"],
               fill=PatternFill("solid", fgColor=VERDE),
               font=Font(color=VERDE_TXT)))
wm.conditional_formatting.add(rango_avance,
    CellIsRule(operator="between", formula=["0.4", "0.7499"],
               fill=PatternFill("solid", fgColor=AMARILLO),
               font=Font(color=AMAR_TXT)))
wm.conditional_formatting.add(rango_avance,
    CellIsRule(operator="lessThan", formula=["0.4"],
               fill=PatternFill("solid", fgColor=ROJO),
               font=Font(color=ROJO_TXT)))

wm.column_dimensions["A"].width = 24
for col in "BCDE":
    wm.column_dimensions[col].width = 16
wm.sheet_view.showGridLines = False

# Grafico de metas: Objetivo vs Ahorrado
chart_metas = BarChart()
chart_metas.type = "col"
chart_metas.title = "Metas: Objetivo vs Ahorrado"
chart_metas.style = 11
chart_metas.y_axis.title = "C$"
data = Reference(wm, min_col=2, max_col=3, min_row=2, max_row=2 + len(metas_demo))
cats = Reference(wm, min_col=1, min_row=3, max_row=2 + len(metas_demo))
chart_metas.add_data(data, titles_from_data=True)
chart_metas.set_categories(cats)
chart_metas.height = 8
chart_metas.width = 18
wm.add_chart(chart_metas, "G2")

# Grafico de avance (% por meta)
chart_av = BarChart()
chart_av.type = "bar"
chart_av.title = "% de avance por meta"
chart_av.style = 10
data2 = Reference(wm, min_col=5, min_row=2, max_row=2 + len(metas_demo))
chart_av.add_data(data2, titles_from_data=True)
chart_av.set_categories(cats)
chart_av.height = 8
chart_av.width = 18
wm.add_chart(chart_av, "G18")

# ===========================================================================
# HOJA 1: DASHBOARD  (se inserta al inicio)
# ===========================================================================
wd = wb.create_sheet("Dashboard", 0)
wd.sheet_view.showGridLines = False
titulo(wd, "A1", "RESUMEN GENERAL  -  FINANZAS PERSONALES", size=18)
wd.merge_cells("A1:G1")
wd.row_dimensions[1].height = 30

# --- Selector de mes / anio ---
wd["A3"] = "Mes a analizar:"
wd["A3"].font = Font(bold=True)
wd["B3"] = 6
wd["A4"] = "Año:"
wd["A4"].font = Font(bold=True)
wd["B4"] = 2026
wd["C3"] = '=CHOOSE(B3,"Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre")'
wd["C3"].font = Font(bold=True, color=AZUL_OSC)
for c in ("B3", "B4"):
    wd[c].fill = PatternFill("solid", fgColor=AZUL_CLARO)
    wd[c].border = BORDE
    wd[c].alignment = Alignment(horizontal="center")

dv_mes = DataValidation(type="whole", operator="between",
                        formula1="1", formula2="12", allow_blank=False)
dv_mes.error = "Ingresa un numero de mes entre 1 y 12."
dv_mes.errorTitle = "Mes invalido"
wd.add_data_validation(dv_mes)
dv_mes.add("B3")

# --- Tabla KPI ---
encabezado(wd["A6"], "Concepto")
encabezado(wd["B6"], "Monto")
kpis = [
    ("Ingresos del mes",
     "=SUMIFS(Ingresos!$C:$C,Ingresos!$F:$F,$B$3,Ingresos!$E:$E,$B$4)"),
    ("Gastos del mes",
     "=SUMIFS(Gastos!$D:$D,Gastos!$G:$G,$B$3,Gastos!$E:$E,$B$4)"),
    ("Ahorro del mes", "=B7-B8"),
    ("Dinero disponible", "=SUM(Ingresos!$C:$C)-SUM(Gastos!$D:$D)"),
    ("Fondo para casa", "=Metas!C3"),
    ("% destinado al ahorro", "=IF(B7=0,0,B9/B7)"),
]
for idx, (concepto, formula) in enumerate(kpis, start=7):
    wd[f"A{idx}"] = concepto
    wd[f"A{idx}"].font = Font(bold=True)
    wd[f"A{idx}"].fill = PatternFill("solid", fgColor=GRIS_CLARO)
    wd[f"A{idx}"].border = BORDE
    wd[f"B{idx}"] = formula
    wd[f"B{idx}"].border = BORDE
    wd[f"B{idx}"].number_format = MONEDA
    wd[f"B{idx}"].font = Font(bold=True)
wd["B12"].number_format = PORC  # % destinado al ahorro

# Semaforo del indicador de ahorro (fila 12)
wd.conditional_formatting.add("B12",
    CellIsRule(operator="greaterThanOrEqual", formula=["0.2"],
               fill=PatternFill("solid", fgColor=VERDE), font=Font(color=VERDE_TXT, bold=True)))
wd.conditional_formatting.add("B12",
    CellIsRule(operator="between", formula=["0.1", "0.1999"],
               fill=PatternFill("solid", fgColor=AMARILLO), font=Font(color=AMAR_TXT, bold=True)))
wd.conditional_formatting.add("B12",
    CellIsRule(operator="lessThan", formula=["0.1"],
               fill=PatternFill("solid", fgColor=ROJO), font=Font(color=ROJO_TXT, bold=True)))

# Estado general del mes (texto + semaforo)
wd["A13"] = "Estado del mes"
wd["A13"].font = Font(bold=True)
wd["A13"].fill = PatternFill("solid", fgColor=GRIS_CLARO)
wd["A13"].border = BORDE
wd["B13"] = ('=IF(B7=0,"Sin ingresos",IF(B8>B7,"ROJO: gastas más de lo que ingresas",'
             'IF(B8>0.9*B7,"AMARILLO: cuidado con los gastos","VERDE: vas bien")))')
wd["B13"].border = BORDE
wd["B13"].font = Font(bold=True)
wd.conditional_formatting.add("B13",
    FormulaRule(formula=['ISNUMBER(SEARCH("VERDE",B13))'],
                fill=PatternFill("solid", fgColor=VERDE), font=Font(color=VERDE_TXT, bold=True)))
wd.conditional_formatting.add("B13",
    FormulaRule(formula=['ISNUMBER(SEARCH("AMARILLO",B13))'],
                fill=PatternFill("solid", fgColor=AMARILLO), font=Font(color=AMAR_TXT, bold=True)))
wd.conditional_formatting.add("B13",
    FormulaRule(formula=['ISNUMBER(SEARCH("ROJO",B13))'],
                fill=PatternFill("solid", fgColor=ROJO), font=Font(color=ROJO_TXT, bold=True)))

# --- Tabla Presupuesto / Gastos por categoria (D6:G14) ---
encabezado(wd["D6"], "Categoría")
encabezado(wd["E6"], "Presupuesto")
encabezado(wd["F6"], "Gastado (mes)")
encabezado(wd["G6"], "% usado")
presupuesto_demo = {
    "Alimentación": 6000, "Transporte": 3000, "Servicios": 2500,
    "Estudios": 2000, "Casa": 4000, "Salud": 2000,
    "Entretenimiento": 1500, "Otros": 1500,
}
for i, cat in enumerate(CATEGORIAS, start=7):
    wd[f"D{i}"] = cat
    wd[f"D{i}"].border = BORDE
    wd[f"E{i}"] = presupuesto_demo[cat]
    wd[f"E{i}"].number_format = MONEDA
    wd[f"E{i}"].border = BORDE
    wd[f"F{i}"] = (f'=SUMIFS(Gastos!$D:$D,Gastos!$B:$B,$D{i},'
                   f'Gastos!$G:$G,$B$3,Gastos!$E:$E,$B$4)')
    wd[f"F{i}"].number_format = MONEDA
    wd[f"F{i}"].border = BORDE
    wd[f"G{i}"] = f'=IF(E{i}=0,0,F{i}/E{i})'
    wd[f"G{i}"].number_format = PORC
    wd[f"G{i}"].border = BORDE

# Semaforo presupuesto (% usado) D..G filas 7-14
rango_pres = "G7:G14"
wd.conditional_formatting.add(rango_pres,
    CellIsRule(operator="greaterThan", formula=["1"],
               fill=PatternFill("solid", fgColor=ROJO), font=Font(color=ROJO_TXT)))
wd.conditional_formatting.add(rango_pres,
    CellIsRule(operator="between", formula=["0.8", "1"],
               fill=PatternFill("solid", fgColor=AMARILLO), font=Font(color=AMAR_TXT)))
wd.conditional_formatting.add(rango_pres,
    CellIsRule(operator="lessThan", formula=["0.8"],
               fill=PatternFill("solid", fgColor=VERDE), font=Font(color=VERDE_TXT)))

# --- Tabla Evolucion mensual (A16:D28) para el anio seleccionado ---
wd["A16"] = "Evolución mensual (año seleccionado)"
wd["A16"].font = Font(bold=True, color=AZUL_OSC, size=12)
encabezado(wd["A17"], "Mes")
encabezado(wd["B17"], "Ingresos")
encabezado(wd["C17"], "Gastos")
encabezado(wd["D17"], "Ahorro")
for k, nombre in enumerate(MESES, start=0):
    fila = 18 + k
    mes_num = k + 1
    wd[f"A{fila}"] = nombre
    wd[f"A{fila}"].border = BORDE
    wd[f"B{fila}"] = (f'=SUMIFS(Ingresos!$C:$C,Ingresos!$F:$F,{mes_num},'
                      f'Ingresos!$E:$E,$B$4)')
    wd[f"C{fila}"] = (f'=SUMIFS(Gastos!$D:$D,Gastos!$G:$G,{mes_num},'
                      f'Gastos!$E:$E,$B$4)')
    wd[f"D{fila}"] = f'=B{fila}-C{fila}'
    for col in "BCD":
        wd[f"{col}{fila}"].number_format = MONEDA
        wd[f"{col}{fila}"].border = BORDE

# Anchos de columna del dashboard
wd.column_dimensions["A"].width = 24
wd.column_dimensions["B"].width = 18
wd.column_dimensions["C"].width = 16
wd.column_dimensions["D"].width = 18
wd.column_dimensions["E"].width = 14
wd.column_dimensions["F"].width = 16
wd.column_dimensions["G"].width = 10

# --- GRAFICOS DEL DASHBOARD ---
# 1) Gastos por categoria (torta)
pie = PieChart()
pie.title = "Gastos por categoría (mes)"
labels = Reference(wd, min_col=4, min_row=7, max_row=14)
vals = Reference(wd, min_col=6, min_row=6, max_row=14)
pie.add_data(vals, titles_from_data=True)
pie.set_categories(labels)
pie.height = 8
pie.width = 12
pie.dataLabels = None
wd.add_chart(pie, "I3")

# 2) Evolucion del ahorro (linea)
line = LineChart()
line.title = "Evolución del ahorro"
line.style = 12
line.y_axis.title = "C$"
data_line = Reference(wd, min_col=4, min_row=17, max_row=29)  # Ahorro + header
cats_line = Reference(wd, min_col=1, min_row=18, max_row=29)
line.add_data(data_line, titles_from_data=True)
line.set_categories(cats_line)
line.height = 8
line.width = 16
wd.add_chart(line, "I20")

# 3) Presupuesto vs gastado (barras)
bar = BarChart()
bar.type = "col"
bar.title = "Presupuesto vs Gastado"
bar.style = 10
data_bar = Reference(wd, min_col=5, max_col=6, min_row=6, max_row=14)
cats_bar = Reference(wd, min_col=4, min_row=7, max_row=14)
bar.add_data(data_bar, titles_from_data=True)
bar.set_categories(cats_bar)
bar.height = 8
bar.width = 16
wd.add_chart(bar, "I37")

# ---------------------------------------------------------------------------
# Guardar
# ---------------------------------------------------------------------------
salida = "/projects/sandbox/Finanzas_Personales.xlsx"
wb.save(salida)
print("Archivo generado:", salida)
print("Hojas:", wb.sheetnames)
